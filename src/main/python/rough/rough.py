# Author: Dahir Muhammad Dahir
# Date: 17th-05-2019
# About: making some basic test, to see if things are working
# as supposed

import openpyxl as excel
from openpyxl.styles import Font
import datetime

workbook = excel.Workbook()
worksheet1 = workbook.active

worksheet1.title = "range names"
bold = Font(bold=True)
dict1 = {"B": "12:30 AM", "A": "Monday", "C": "2019", "E": "Test"}
space = {"A": " ", "B": " ", "C": " "}
values = list(dict1.values())

for i in range(3):
    worksheet1.append(dict1)
"""
i = 0
for row in worksheet1.iter_rows(min_row=1, max_row=1, max_col=len(dict1.values())):
    for cell in row:
        cell.font = bold
        cell.value = values[i]
        i += 1

for row in range(1, 4):
    worksheet1.append(space)

i = 0
for row in worksheet1.iter_rows(min_row=1, max_row=1, max_col=len(dict1.values())):
    for cell in row:
        cell.font = bold
        cell.value = values[i]
        i += 1

        # cell.font = bold
        # worksheet1[cell] = "time"
# worksheet2 = workbook.create_sheet("Pi")
# worksheet2["F5"] = 3.14

# worksheet3 = workbook.create_sheet("Data")

# for row in range(10, 20):
#    for col in range(27, 54):
#        _ = worksheet3.cell(column=col, row=row, value="{}".format(get_column_letter(col)))
"""
workbook.save("sample.xlsx")


def add_table_field(self, field):
    for dict_item in self.all_table_fields:
        if self.all_table_fields[dict_item] == field:
            return
    key = self.get_unique_key()
    self.all_table_fields[key] = field
